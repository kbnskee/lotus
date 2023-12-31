import os, sys
from pathlib import Path
from django.core.management.base import BaseCommand
from django.core.management import call_command
import openpyxl
from django.apps import apps
from lotus.conf import BASE_DATA

# import gspread
# from google.auth import exceptions
# from google.oauth2 import service_account
# from openpyxl import Workbook

from lotus.models import (
    User,
    App,
    Page,
    Group,
    GroupApp,
    GroupPage,
    UserGroup,
    Gender,
    ContactType,
    AddressType,
    FamilyRelationshipType,
    BloodType,
    Suffix,
    Citizenship,
    CivilStatus,
    Role,
    ActivityLog,
    ActivityType
)

base_dir=BASE_DIR = Path(__file__).resolve().parent.parent.parent.parent

predefined="predefined"
predefined_lotus=predefined+'_lotus.xlsx'
predefined_app=predefined+'_app.xlsx'

msg_exiting="Exiting!"

    


class Command(BaseCommand):
    help = "kadari"

    def add_arguments(self, parser):
        parser.add_argument("arg", type=str)

    def handle(self, *args, **kwargs):
        _arg=kwargs['arg']

        if _arg=="info":
            self.__info()

        elif _arg=="lotus-init":
            self.__init_lotus()

        elif _arg=="init-app":
            self.__init_app()

        # elif _arg=="lotus-pull-predefined":
        #     self.__get_lotus_from_googleapis()

        elif _arg=="supload":
            from apps.subjectms.models import Subject
            excel_data = openpyxl.load_workbook("supload.xlsx")
            sheet=excel_data["Sheet1"]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                Subject.objects.create(
                    id=row['id'],
                    name=row['name'],
                    description=row['description'],
                    is_enabled=row['is_enabled'],
                    created_by_id=row['created_by_id'],
                    type_id=row['type_id'],
                    code=row['code'],
                    passing_score_percentage=row['passing_score_percentage'],
                    passing_validity=row['passing_validity'],
                    written_works_percentage=row['written_works_percentage'],
                    practical_test_percentage=row['practical_test_percentage'],
                    quarterly_exam_percentage=row['quarterly_exam_percentage'],
                    year_level_id=row['year_level_id']
                )
        elif _arg=="clearbasedata":
            self.__clear_tables()

        elif _arg=="lotus-reset":
            self.__reset()

        elif _arg=="factoryreset":
            self.__factory_reset()

        elif _arg=="show-excel-models":
            self.__get_excel_models()
                
        elif _arg=="createsuperuser":
            self.__create_superuser()

        elif _arg=="upload_students":
            self.__upload_student()


    def __info(self):
        self.stdout.write(f"kadari 1.0")

    def __init_lotus(self):
        self.__delete_table_contents()
        self.__create_superuser()
        self.__create_superuser_group()
        self.__create_suuperuser_usergroup()
        self.__create_lotus_app()
        self.__create_lotus_pages()

        


    def __init_app(self):
        # self.__delete_app_table_contnet()
        self.__upload_excel()

    def __reset(self):
        self.stdout.write("1. Resetting lotus...")
        self.__delete_table_contents()
        self.stdout.write("1. Resetting lotus...OK!")
        sys.exit()

    def __delete_table_contents(self):
        self.stdout.write("1. Deleting Lotus tables...")
        User.objects.all().delete()
        App.objects.all().delete()
        Page.objects.all().delete()
        Group.objects.all().delete()
        UserGroup.objects.all().delete()
        GroupApp.objects.all().delete()
        GroupPage.objects.all().delete()
        ContactType.objects.all().delete()
        Gender.objects.all().delete()
        AddressType.objects.all().delete()
        FamilyRelationshipType.objects.all().delete()
        BloodType.objects.all().delete()
        Suffix.objects.all().delete()
        Citizenship.objects.all().delete()
        CivilStatus.objects.all().delete()
        Role.objects.all().delete()
        ActivityType.objects.all().delete()
        ActivityLog.objects.all().delete()
        self.stdout.write("1. Deleting Lotus tables...OK!")

    def __delete_app_table_contnet(self):
        from apps.schoolms.models import School, Stage, YearLevel
        School.objects.all().delete()
        Stage.objects.all().delete()
        YearLevel.objects.all().delete()
        
        self.stdout.write("Cleaning School, Stage, YearLevel...OK!")


    def __create_superuser(self):
        self.stdout.write("2. Creating createsuperuser...")
        user=User(id=1,username="kadari")
        user.set_password("kadari")
        user.save()
        self.stdout.write("2. Creating createsuperuser...OK!")

    def __create_superuser_group(self):
        self.stdout.write("3. Createing Superuser Group...")
        group=Group(id=1,name="superuser")
        group.created_by=self.__get_superuser()
        group.save()
        self.stdout.write("3. Createing Superuser Group...OK!")

    def __get_superuser_group(self):
        return Group.objects.get(name="superuser")
    
    def __get_superuser(self):
        return User.objects.get(username="kadari")
    
    def __create_suuperuser_usergroup(self):
        self.stdout.write("4. creating createsuperuser...")
        user_group=UserGroup.objects.create(id=1,user=self.__get_superuser(),
                                            group=self.__get_superuser_group(),
                                            created_by=self.__get_superuser())
        user_group.save()
        self.stdout.write("4. creating createsuperuser... OK")



    def __create_lotus_app(self):
        self.stdout.write("5. Creating Lotus App...")
        App.objects.create(id=100,name="lotus",created_by_id=self.__get_superuser().id,is_enabled=True)
        GroupApp.objects.create(id=1,group_id=1,app_id=100)
        self.stdout.write("5. Creating Lotus App... OK")
    

    def __get_lotus_app(self):
        return App.objects.get(id=100,name="lotus")
    

    # def __get_lotus_from_googleapis(self):
    #     scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]

    #     try:
    #         credentials = service_account.Credentials.from_service_account_file(
    #             "lotus-dio-predat-e366e06d8b99.json", scopes=scope
    #         )
    #     except exceptions.DefaultCredentialsError:
    #         raise ValueError("Failed to authenticate with Google Sheets API. Check your credentials.")

    #     gc = gspread.authorize(credentials)

    #     # Open the Google Sheet using its title or URL
    #     # Replace 'Your Google Sheet Name' with your actual sheet name or URL
    #     workbook = gc.open("https://docs.google.com/spreadsheets/d/1S9mmEmpdYGYcrhoc2Xi9n-U2eDV-OQbC7qA3ySwCZTg/edit")

    #     # Create a new Excel workbook
    #     output_excel_file = "predefined_lotus.xlsx"
    #     wb = Workbook()

    #     # Iterate through all sheets in the workbook
    #     for sheet in workbook:
    #         # Get all values from the sheet
    #         data = sheet.get_all_values()

    #         # Add a new worksheet to the Excel workbook
    #         ws = wb.create_sheet(title=sheet.title)

    #         # Write the data to the worksheet
    #         for row in data:
    #             ws.append(row)

    #     # Remove the default empty sheet created by openpyxl
    #     if 'Sheet' in wb.sheetnames:
    #         wb.remove(wb['Sheet'])

    #     # Save the entire workbook to a local Excel file
    #     wb.save(output_excel_file)

    #     print(f"Workbook downloaded and saved to '{output_excel_file}'")

    
    def __create_lotus_pages(self):
        self.stdout.write("6. Creating Lotus Pages...")
        # app=self.__get_lotus_app()
        # Page.objects.create(id=str(app.id)+"001",app=app,path="lotus_app_list",name="lotus_app_list",is_enabled=True,created_by=self.__get_superuser())
        # GroupPage.objects.create(group_id=1,page_id=str(app.id)+"001")
        # Page.objects.create(id=str(app.id)+"002",app=app,path="lotus_app_add",name="lotus_app_add",is_enabled=True,created_by=self.__get_superuser())
        # GroupPage.objects.create(group_id=1,page_id=str(app.id)+"002")
        # Page.objects.create(id=str(app.id)+"003",app=app,path="lotus_app_details",name="lotus_app_details",is_enabled=True,created_by=self.__get_superuser())
        # GroupPage.objects.create(group_id=1,page_id=str(app.id)+"003")
        # Page.objects.create(id=str(app.id)+"004",app=app,path="lotus_app_update",name="lotus_app_update",is_enabled=True,created_by=self.__get_superuser())
        # GroupPage.objects.create(group_id=1,page_id=str(app.id)+"004")
        # Page.objects.create(id=str(app.id)+"005",app=app,path="lotus_app_delete",name="lotus_app_delete",is_enabled=True,created_by=self.__get_superuser())
        # GroupPage.objects.create(group_id=1,page_id=str(app.id)+"005")
        # Page.objects.create(id=str(app.id)+"006",app=app,path="lotus_page_list",name="lotus_page_list",is_enabled=True,created_by=self.__get_superuser())
        # GroupPage.objects.create(group_id=1,page_id=str(app.id)+"006")
        # Page.objects.create(id=str(app.id)+"007",app=app,path="lotus_page_add",name="lotus_page_add",is_enabled=True,created_by=self.__get_superuser())
        # GroupPage.objects.create(group_id=1,page_id=str(app.id)+"007")
        # Page.objects.create(id=str(app.id)+"008",app=app,path="lotus_page_details",name="lotus_page_details",is_enabled=True,created_by=self.__get_superuser())
        # GroupPage.objects.create(group_id=1,page_id=str(app.id)+"008")
        # Page.objects.create(id=str(app.id)+"009",app=app,path="lotus_page_update",name="lotus_page_update",is_enabled=True,created_by=self.__get_superuser())
        # GroupPage.objects.create(group_id=1,page_id=str(app.id)+"009")
        # Page.objects.create(id=str(app.id)+"010",app=app,path="lotus_page_delete",name="lotus_page_delete",is_enabled=True,created_by=self.__get_superuser())
        # GroupPage.objects.create(group_id=1,page_id=str(app.id)+"010")
        # Page.objects.create(id=str(app.id)+"011",app=app,path="lotus_group_list",name="lotus_group_list",is_enabled=True,created_by=self.__get_superuser())
        # GroupPage.objects.create(group_id=1,page_id=str(app.id)+"011")
        # Page.objects.create(id=str(app.id)+"012",app=app,path="lotus_group_add",name="lotus_group_add",is_enabled=True,created_by=self.__get_superuser())
        # GroupPage.objects.create(group_id=1,page_id=str(app.id)+"012")
        # Page.objects.create(id=str(app.id)+"013",app=app,path="lotus_group_details",name="lotus_group_details",is_enabled=True,created_by=self.__get_superuser())
        # GroupPage.objects.create(group_id=1,page_id=str(app.id)+"013")
        # Page.objects.create(id=str(app.id)+"014",app=app,path="lotus_group_app_add",name="lotus_group_app_add",is_enabled=True,created_by=self.__get_superuser())
        # GroupPage.objects.create(group_id=1,page_id=str(app.id)+"014")
        # Page.objects.create(id=str(app.id)+"015",app=app,path="lotus_group_page_add",name="lotus_group_page_add",is_enabled=True,created_by=self.__get_superuser())
        # GroupPage.objects.create(group_id=1,page_id=str(app.id)+"015")
        # Page.objects.create(id=str(app.id)+"016",app=app,path="lotus_group_app_delete",name="lotus_group_app_delete",is_enabled=True,created_by=self.__get_superuser())
        # GroupPage.objects.create(group_id=1,page_id=str(app.id)+"016")
        # Page.objects.create(id=str(app.id)+"017",app=app,path="lotus_group_update",name="lotus_group_update",is_enabled=True,created_by=self.__get_superuser())
        # GroupPage.objects.create(group_id=1,page_id=str(app.id)+"017")
        # Page.objects.create(id=str(app.id)+"018",app=app,path="lotus_group_delete",name="lotus_group_delete",is_enabled=True,created_by=self.__get_superuser())
        # GroupPage.objects.create(group_id=1,page_id=str(app.id)+"018")
        # Page.objects.create(id=str(app.id)+"019",app=app,path="lotus_user_list",name="lotus_user_list",is_enabled=True,created_by=self.__get_superuser())
        # GroupPage.objects.create(group_id=1,page_id=str(app.id)+"019")
        # Page.objects.create(id=str(app.id)+"020",app=app,path="lotus_user_employee_details",name="lotus_user_employee_details",is_enabled=True,created_by=self.__get_superuser())
        # GroupPage.objects.create(group_id=1,page_id=str(app.id)+"020")
        # Page.objects.create(id=str(app.id)+"021",app=app,path="lotus_user_student_details",name="lotus_user_student_details",is_enabled=True,created_by=self.__get_superuser())
        # GroupPage.objects.create(group_id=1,page_id=str(app.id)+"021")
        # Page.objects.create(id=str(app.id)+"022",app=app,path="lotus_user_update",name="lotus_user_update",is_enabled=True,created_by=self.__get_superuser())
        # GroupPage.objects.create(group_id=1,page_id=str(app.id)+"022")
        # Page.objects.create(id=str(app.id)+"023",app=app,path="lotus_user_delete",name="lotus_user_delete",is_enabled=True,created_by=self.__get_superuser())
        # GroupPage.objects.create(group_id=1,page_id=str(app.id)+"023")
        # Page.objects.create(id=str(app.id)+"024",app=app,path="lotus_importer",name="lotus_importer",is_enabled=True,created_by=self.__get_superuser())
        # GroupPage.objects.create(group_id=1,page_id=str(app.id)+"024")
        # Page.objects.create(id=str(app.id)+"025",app=app,path="lotus_404",name="lotus_404",is_enabled=True,created_by=self.__get_superuser())
        # GroupPage.objects.create(group_id=1,page_id=str(app.id)+"025")
        excel_file_path = predefined_lotus

        try:
            workbook = openpyxl.load_workbook(excel_file_path, read_only=True, data_only=True)

            for sheet in workbook.sheetnames:
                django_model = globals().get(sheet, None)
                self.stdout.write(f"KDR: initializing django model: {django_model}")

                if django_model:
                    sheet_data = workbook[sheet]

                    headers = [cell.value for cell in sheet_data[1]]
                    for row in sheet_data.iter_rows(min_row=2, values_only=True):
                        data = dict(zip(headers, row))
                        instance = django_model()
                        for field, value in data.items():
                            setattr(instance, field, value)
                        instance.save()
                    
                else:
                    self.stdout.write(f'Model not found for sheet {sheet}. Check your models and sheet names.')
                    sys.exit()
                self.stdout.write(f"KDR: initializing django model: {django_model}...OK!")
        except Exception as e:
            self.stdout.write(f'Error: {str(e)}')
            self.stdout.write(f"{msg_exiting}...")
            sys.exit()
        
        sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
        self.stdout.write("6. Creating Lotus Pages...OK!")




    def __upload_excel(self):
        # self.stdout.write("KDR: init-app. Upload Excel...")
        # excel_file_path = predefined_app

        # try:
        #     workbook = openpyxl.load_workbook(excel_file_path, read_only=True, data_only=True)

        #     for sheet in workbook.sheetnames:
        #         django_model = globals().get(sheet, None)
        #         self.stdout.write(f"KDR: initializing django model: {django_model}")

        #         if django_model:
        #             sheet_data = workbook[sheet]

        #             headers = [cell.value for cell in sheet_data[1]]
        #             for row in sheet_data.iter_rows(min_row=2, values_only=True):
        #                 data = dict(zip(headers, row))
        #                 instance = django_model()
        #                 for field, value in data.items():
        #                     setattr(instance, field, value)
        #                 instance.save()
                    
        #         else:
        #             self.stdout.write(f'Model not found for sheet {sheet}. Check your models and sheet names.')
        #             sys.exit()
        #         self.stdout.write(f"KDR: initializing django model: {django_model}...OK!")
        # except Exception as e:
        #     self.stdout.write(f'Error: {str(e)}')
        #     self.stdout.write(f"{msg_exiting}...")
        
        # sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
        
        
        from apps.schoolms.apps import SchoolmsConfig
        from apps.companybasems.apps import CompanybasemsConfig
        from apps.tuitionms.apps import TuitionmsConfig
        from apps.teacher.apps import TeacherConfig
        from apps.subjectms.apps import SubjectmsConfig
        from apps.sectionms.apps import SectionmsConfig
        from apps.itadmin.apps import ItadminConfig
        from apps.gradems.apps import GrademsConfig
        from apps.conductms.apps import ConductmsConfig
        from apps.cashier.apps import CashierConfig


        # self.__upload_app_models(SchoolmsConfig.name)
        # self.__upload_app_models(CompanybasemsConfig.name)
        # self.__upload_app_models(TuitionmsConfig.name)
        # self.__upload_app_models(TeacherConfig.name)
        # self.__upload_app_models(SubjectmsConfig.name)
        # self.__upload_app_models(SectionmsConfig.name)
        # self.__upload_app_models(ItadminConfig.name)
        # self.__upload_app_models(GrademsConfig.name)
        # self.__upload_app_models(ConductmsConfig.name)
        # self.__upload_app_models(CashierConfig.name)


    def __upload_app_models(self,model):
        print(f"KDR: init-app. Upload Excel... {model} ")
        excel_file_path = os.path.join(os.getcwd(), predefined+"_"+model.split(".")[1]+".xlsx")

        try:
            excel_file_path = excel_file_path
            workbook = openpyxl.load_workbook(excel_file_path, read_only=True, data_only=True)

            for sheet in workbook.sheetnames:
                model_name = sheet.capitalize()
                django_model = apps.get_model(app_label=model.split(".")[1], model_name=model_name)
                
                print(f"KDR: initializing django model: {django_model}")
                if django_model:
                    sheet_data = workbook[sheet]
                    headers = [cell.value for cell in sheet_data[1]]
                    for row in sheet_data.iter_rows(min_row=2, values_only=True):
                        data = dict(zip(headers, row))
                        instance = django_model(**data)
                        instance.save()
                    print(f"KDR: initializing django model: {django_model}...OK!")
                else:
                    print(f'Model not found for sheet {sheet}. Check your models and sheet names.')
                    sys.exit()

        except Exception as e:
            print(f'Error: {str(e)}')
            self.stdout.write(f"{msg_exiting}...")


    def __upload_student(self):
        print(f"KDR: init-app. Upload Excel...  ")
        excel_file_path = os.path.join(os.getcwd(), "students.xlsx")

        try:
            excel_file_path = excel_file_path
            workbook = openpyxl.load_workbook(excel_file_path, read_only=True, data_only=True)

            for sheet in workbook.sheetnames:
                model_name = sheet.capitalize()
                django_model = apps.get_model(app_label="students", model_name=model_name)
                
                print(f"KDR: initializing django model: {django_model}")
                if django_model:
                    sheet_data = workbook[sheet]
                    headers = [cell.value for cell in sheet_data[1]]
                    for row in sheet_data.iter_rows(min_row=2, values_only=True):
                        data = dict(zip(headers, row))
                        instance = django_model(**data)
                        instance.save()
                    print(f"KDR: initializing django model: {django_model}...OK!")
                else:
                    print(f'Model not found for sheet {sheet}. Check your models and sheet names.')
                    sys.exit()

        except Exception as e:
            print(f'Error: {str(e)}')
            self.stdout.write(f"{msg_exiting}...")
        

    def __get_excel_models(self):
        excel_file_path = predefined_app
        self.stdout.write("Generating List from predefined.xlsx...")
        try:
            workbook = openpyxl.load_workbook(excel_file_path, read_only=True, data_only=True)
            for sheet in workbook.sheetnames:
                django_model = globals().get(sheet, None)
                self.stdout.write(f"django model: {django_model}")
            self.stdout.write("Generating List from predefined.xlsx...OK!")
        except Exception as e:
            self.stdout.write(f'Error: {str(e)}...{msg_exiting}')
            sys.exit()
        
    
    def __factory_reset(self):
        user_input = input("WARNING!!! This will reset the whole application. Are you sure you want to continue? Yes/No: ")
        if user_input.lower() == "no":
            self.stdout.write(f"{msg_exiting}...")
            sys.exit()
        elif user_input.lower() == "yes":
            final = input("Man, this is going to delete all your data! Are you sure? Yes/No: ")
            if final.lower()=="no":
                self.stdout.write(f"{msg_exiting}...")
                sys.exit()
            elif final.lower()=="yes":
                self.stdout.write("Okay then...")
                self.stdout.write("resetting...")
                User.objects.all().delete()
                App.objects.all().delete()
                Group.objects.all().delete()
                self.stdout.write("resetting...OK")
            else:
                self.stdout.write(f"{msg_exiting}...")
                sys.exit()
        else:
            self.stdout.write(f"{msg_exiting}...")
            sys.exit()
    
    

    def __clear_tables(self):
        answer=input(f"This command clears all Role, App, AppUser, Page, PageUser in your database. Do you want to continue? Yes or No. Default=Yes... ")
        if answer=="Yes" or not answer.strip():
            App.objects.all().delete()
            Page.objects.all().delete()
            Group.objects.all().delete()
            UserGroup.objects.all().delete()
            self.stdout.write(f"Clearing tables... Ok")
        else:
            self.stdout.write(f"loadbasedata... Cancelled")


    
            